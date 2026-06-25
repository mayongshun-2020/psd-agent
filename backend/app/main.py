from __future__ import annotations

import json
import shutil
import uuid
from pathlib import Path

from fastapi import FastAPI, File, Form, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse
from pydantic import ValidationError

from .defaults import load_workflow_defaults
from .models import UploadedAsset, WorkflowArtifacts, WorkflowRequest, WorkflowResult
from .pipeline import classify_asset, run_pipeline
from .render import build_design_spec, write_artifacts

APP_ROOT = Path(__file__).resolve().parents[1]
RUNS_ROOT = APP_ROOT / "runs"

app = FastAPI(title="BrandOS AI Design Platform", version="0.3.0")
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/api/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.get("/api/config/defaults")
def config_defaults() -> dict[str, object]:
    defaults = load_workflow_defaults()
    return {
        "payload": defaults,
        "prompts": defaults["prompts"],
        "workflowModes": ["smart_recommend", "strict_brand"],
        "outputTypes": ["detail_page", "figma_page", "psd_file", "main_image", "banner"],
        "stages": [
            {"id": "product_understanding", "title": "商品理解 Agent", "icon": "eye"},
            {"id": "product_brief", "title": "Product Brief", "icon": "layers"},
            {"id": "brand_knowledge", "title": "品牌知识库 / 规则版本", "icon": "library"},
            {"id": "page_planner", "title": "页面规划 Agent", "icon": "palette"},
            {"id": "layout_engine", "title": "Layout Engine", "icon": "grid"},
            {"id": "copy", "title": "文案 Agent", "icon": "type"},
            {"id": "figma_psd", "title": "Figma / PSD 生成 Agent", "icon": "file-image"},
            {"id": "design_score", "title": "Design Score", "icon": "check-circle"},
            {"id": "output_review", "title": "输出、审核与反馈", "icon": "check-circle"},
        ],
    }


def _safe_filename(name: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in "._-()[] " else "_" for ch in name)
    return cleaned.strip() or "asset"


def _extract_spreadsheet_text(path: Path) -> str | None:
    if path.suffix.lower() not in {".xlsx", ".xlsm"}:
        return None
    try:
        from openpyxl import load_workbook
    except Exception:
        return None

    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        lines: list[str] = []
        for sheet in workbook.worksheets[:6]:
            lines.append(f"[Sheet] {sheet.title}")
            count = 0
            for row in sheet.iter_rows(values_only=True):
                values = [str(value).strip() for value in row if value not in (None, "")]
                if values:
                    lines.append(" | ".join(values))
                    count += 1
                if count >= 120:
                    break
        return "\n".join(lines)[:16000]
    except Exception:
        return None


async def _save_assets(files: list[UploadFile], input_dir: Path) -> list[UploadedAsset]:
    input_dir.mkdir(parents=True, exist_ok=True)
    assets: list[UploadedAsset] = []
    for file in files:
        filename = _safe_filename(file.filename or "asset")
        target = input_dir / filename
        with target.open("wb") as buffer:
            shutil.copyfileobj(file.file, buffer)
        assets.append(
            UploadedAsset(
                name=filename,
                content_type=file.content_type,
                size=target.stat().st_size,
                saved_path=str(target),
                extracted_text=_extract_spreadsheet_text(target),
                bucket=classify_asset(filename, file.content_type),
            )
        )
    return assets


def _merge_payload(incoming: dict) -> dict:
    data = load_workflow_defaults()
    for key, value in incoming.items():
        if isinstance(data.get(key), dict) and isinstance(value, dict):
            merged = dict(data[key])
            merged.update(value)
            data[key] = merged
        else:
            data[key] = value
    return data


@app.post("/api/workflows/generate", response_model=WorkflowResult)
async def generate_workflow(
    payload: str = Form(...),
    files: list[UploadFile] = File(default=[]),
) -> WorkflowResult:
    try:
        incoming = json.loads(payload)
        if not isinstance(incoming, dict):
            raise ValueError("payload 必须是 JSON 对象")
        request = WorkflowRequest.model_validate(_merge_payload(incoming))
    except (json.JSONDecodeError, ValidationError, ValueError) as exc:
        raise HTTPException(status_code=422, detail=str(exc)) from exc

    run_id = uuid.uuid4().hex
    run_dir = RUNS_ROOT / run_id
    input_dir = run_dir / "inputs"
    output_dir = run_dir / "outputs"
    assets = await _save_assets(files, input_dir)

    spreadsheet_text = "\n\n".join(
        asset.extracted_text or "" for asset in assets if asset.extracted_text
    ).strip()
    if spreadsheet_text and spreadsheet_text not in request.product_brief:
        request.product_brief = "\n\n".join(
            part for part in [request.product_brief, spreadsheet_text] if part
        )

    stages, ctx = run_pipeline(request, assets)
    spec = build_design_spec(ctx)
    artifact_paths = write_artifacts(output_dir, spec)

    used_model = any(stage.used_model for stage in stages)
    agent_report = "\n\n".join(ctx.report_parts)

    return WorkflowResult(
        run_id=run_id,
        status="completed" if used_model else "fallback_completed",
        summary="BrandOS 设计任务已完成：含品牌规则分层、页面结构、SVG 预览、Figma/PSD 结构说明、评分与反馈清单。",
        used_deepagents=used_model,
        stages=stages,
        agent_report=agent_report,
        design_spec=spec,
        artifacts=WorkflowArtifacts(
            run_id=run_id,
            output_dir=str(output_dir),
            **artifact_paths,
        ),
        assets=assets,
        warnings=ctx.warnings,
    )


@app.get("/api/workflows/{run_id}/artifacts/{name}")
def download_artifact(run_id: str, name: str) -> FileResponse:
    allowed = {
        "preview.svg": "image/svg+xml",
        "design_spec.json": "application/json",
        "create_detail_page.jsx": "text/plain",
        "README.md": "text/markdown",
    }
    if name not in allowed:
        raise HTTPException(status_code=404, detail="artifact not found")
    path = RUNS_ROOT / run_id / "outputs" / name
    if not path.is_file():
        raise HTTPException(status_code=404, detail="artifact not found")
    return FileResponse(path, media_type=allowed[name], filename=name)
